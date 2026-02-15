#!/usr/bin/env python3
from __future__ import annotations

import html
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


AMENDMENT_DATE = date(2021, 1, 11)
SOURCE_FILES = {
    "4414 Deeds": "Deschutes Heights 4414 Deeds.md",
    "Phase 1 File": "Deschutes Heights Phase 1 Deeds.md",
    "Phase 2 File": "Deschutes Heights Phase 2 Deeds.md",
}


@dataclass
class MatchResult:
    matched_index: int | None
    delta_days: int | None


def parse_us_date(value: str | None) -> date | None:
    if not value:
        return None
    text = value.strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def date_to_str(value: date | None) -> str:
    return value.strftime("%m/%d/%Y") if value else ""


def normalize_space(text: str | None) -> str:
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip()


def strip_tags(text: str) -> str:
    no_tags = re.sub(r"<[^>]+>", " ", text)
    return normalize_space(html.unescape(no_tags))


def normalize_name(text: str | None) -> str:
    s = normalize_space(text).lower()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(
        r"\b(llc|l l c|inc|corp|corporation|co|company|ltd|limited|trustee|trust)\b",
        " ",
        s,
    )
    return normalize_space(s)


def is_lotus_house(name: str | None) -> bool:
    normalized = normalize_name(name)
    return "lotus house development" in normalized


def is_bad_name(name: str | None) -> bool:
    n = normalize_space(name)
    if not n or n in {".", "-", "--"}:
        return True
    lowered = n.lower()
    bad_fragments = [
        "for and in consideration",
        "hereinafter",
        "the following described",
        "situated in the county",
        "mailing address of",
    ]
    if any(fragment in lowered for fragment in bad_fragments):
        return True
    if lowered.startswith(","):
        return True
    letters = len(re.findall(r"[A-Za-z]", n))
    return letters < 3


def parse_lot(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.isdigit():
        return int(text)
    m = re.search(r"\d+", text)
    if m:
        return int(m.group(0))
    return None


def parse_resale_reports(data_dir: Path) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    files = sorted(data_dir.glob("*Homeowner Resale Report.md"))

    for path in files:
        content = path.read_text(encoding="utf-8", errors="ignore")
        period_match = re.search(
            r"Escrow Date[:\s*]*\*?\*?\s*(\d+/\d+/\d+)\s*-\s*(\d+/\d+/\d+)",
            content,
            flags=re.IGNORECASE,
        )
        report_period = (
            f"{period_match.group(1)}-{period_match.group(2)}" if period_match else ""
        )

        for tbody in re.findall(r"<tbody>([\s\S]*?)</tbody>", content, flags=re.IGNORECASE):
            if "No data rows" in tbody:
                continue
            for tr in re.findall(r"<tr>([\s\S]*?)</tr>", tbody, flags=re.IGNORECASE):
                cells = re.findall(r"<t[dh]>([\s\S]*?)</t[dh]>", tr, flags=re.IGNORECASE)
                if len(cells) < 7:
                    continue
                parsed_cells = [strip_tags(cell) for cell in cells]
                lot = parse_lot(parsed_cells[3].lstrip("0") or parsed_cells[3])
                entries.append(
                    {
                        "accountNumber": parsed_cells[0],
                        "newOwner": parsed_cells[1],
                        "address": parsed_cells[2],
                        "lot": lot if lot is not None else 0,
                        "previousOwner": parsed_cells[4],
                        "processDate": parsed_cells[5],
                        "escrowDate": parsed_cells[6],
                        "sourceFile": path.name,
                        "reportPeriod": report_period,
                        "escrowDateObj": parse_us_date(parsed_cells[6]),
                        "processDateObj": parse_us_date(parsed_cells[5]),
                    }
                )

    deduped: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()
    for row in sorted(
        entries,
        key=lambda x: (
            x["lot"],
            x["escrowDateObj"] or date.min,
            normalize_name(x["newOwner"]),
            normalize_name(x["previousOwner"]),
        ),
    ):
        key = (
            row["lot"],
            normalize_name(row["newOwner"]),
            normalize_name(row["previousOwner"]),
            row["escrowDate"],
            normalize_space(row["address"]).lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def match_by_lot(
    left_items: list[dict[str, Any]],
    right_items: list[dict[str, Any]],
    left_date_key: str,
    right_date_key: str,
) -> dict[int, MatchResult]:
    result: dict[int, MatchResult] = {}
    left_by_lot: dict[int, list[dict[str, Any]]] = defaultdict(list)
    right_by_lot: dict[int, list[dict[str, Any]]] = defaultdict(list)

    for item in left_items:
        lot = item.get("lotInt")
        if lot is not None:
            left_by_lot[lot].append(item)
    for item in right_items:
        lot = item.get("lotInt")
        if lot is not None:
            right_by_lot[lot].append(item)

    for lot, lot_left in left_by_lot.items():
        lot_right = right_by_lot.get(lot, [])
        if not lot_right:
            for l in lot_left:
                result[l["_idx"]] = MatchResult(None, None)
            continue

        candidates: list[tuple[int, int, int]] = []
        for l in lot_left:
            l_date = l.get(left_date_key)
            if not isinstance(l_date, date):
                continue
            for r in lot_right:
                r_date = r.get(right_date_key)
                if not isinstance(r_date, date):
                    continue
                delta = abs((l_date - r_date).days)
                candidates.append((delta, l["_idx"], r["_idx"]))

        candidates.sort()
        used_left: set[int] = set()
        used_right: set[int] = set()
        for delta, left_idx, right_idx in candidates:
            if left_idx in used_left or right_idx in used_right:
                continue
            used_left.add(left_idx)
            used_right.add(right_idx)
            result[left_idx] = MatchResult(right_idx, delta)

        for l in lot_left:
            if l["_idx"] not in result:
                result[l["_idx"]] = MatchResult(None, None)

    for item in left_items:
        if item["_idx"] not in result:
            result[item["_idx"]] = MatchResult(None, None)

    return result


def infer_gl_unit(row: dict[str, Any]) -> int | None:
    if row.get("unit") is not None:
        try:
            return int(row["unit"])
        except (TypeError, ValueError):
            pass
    description = normalize_space(row.get("description"))
    m = re.search(r"\bunit\s*#?\s*0*(\d{1,3})\b", description, flags=re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def normalize_year(raw: str) -> int:
    y = int(raw)
    if y < 100:
        return 2000 + y if y < 50 else 1900 + y
    return y


def parse_pages(file_path: Path) -> dict[int, str]:
    content = file_path.read_text(encoding="utf-8", errors="ignore")
    matches = list(re.finditer(r"<!-- PAGE (\d+) -->", content))
    pages: dict[int, str] = {}
    for i, m in enumerate(matches):
        page_num = int(m.group(1))
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(content)
        pages[page_num] = content[start:end]
    return pages


def extract_dated_line(text: str) -> date | None:
    month_map = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }

    m = re.search(r"\bDated[:\s]+(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{2,4})\b", text, flags=re.IGNORECASE)
    if m:
        mm, dd, yy = int(m.group(1)), int(m.group(2)), normalize_year(m.group(3))
        try:
            return date(yy, mm, dd)
        except ValueError:
            pass

    m = re.search(r"\bDated[:\s]+([A-Za-z]+)\s+(\d{1,2}),?\s+(\d{2,4})\b", text, flags=re.IGNORECASE)
    if m:
        month = month_map.get(m.group(1).lower())
        if month:
            day, year = int(m.group(2)), normalize_year(m.group(3))
            try:
                return date(year, month, day)
            except ValueError:
                pass

    m = re.search(
        r"\bDated\s+this\s+(\d{1,2})(?:st|nd|rd|th)?\s+day\s+of\s+([A-Za-z]+),?\s+(\d{4})\b",
        text,
        flags=re.IGNORECASE,
    )
    if m:
        day, month_name, year = int(m.group(1)), m.group(2).lower(), int(m.group(3))
        month = month_map.get(month_name)
        if month:
            try:
                return date(year, month, day)
            except ValueError:
                pass
    return None


def extract_recording_stamp(text: str) -> date | None:
    m = re.search(
        r"\b(\d{1,2})/(\d{1,2})/(\d{4})\s+\d{1,2}:\d{2}\s*(?:AM|PM)\b",
        text,
        flags=re.IGNORECASE,
    )
    if m:
        mm, dd, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(yy, mm, dd)
        except ValueError:
            return None
    return None


def auto_fit_columns(ws, min_width: int = 10, max_width: int = 60) -> None:
    for column in ws.columns:
        values = [normalize_space(str(cell.value)) for cell in column if cell.value is not None]
        if not values:
            continue
        width = max(min_width, min(max_width, max(len(v) for v in values) + 2))
        ws.column_dimensions[column[0].column_letter].width = width


def currency(value: float) -> float:
    return round(float(value), 2)


def main() -> None:
    project_root = Path(__file__).resolve().parents[1]
    data_dir = project_root / "data"
    working_dir = project_root / "working"
    working_dir.mkdir(exist_ok=True)

    deeds_path = working_dir / "deeds_updated.json"
    if not deeds_path.exists():
        deeds_path = project_root / "all_deeds_final.json"
    gl_path = project_root / "gl_complete.json"

    deeds: list[dict[str, Any]] = json.loads(deeds_path.read_text(encoding="utf-8"))
    gl_entries: list[dict[str, Any]] = json.loads(gl_path.read_text(encoding="utf-8"))
    resale_entries = parse_resale_reports(data_dir)

    for idx, row in enumerate(resale_entries):
        row["_idx"] = idx
        row["lotInt"] = parse_lot(row.get("lot"))
        row["newOwner"] = normalize_space(row.get("newOwner"))
        row["previousOwner"] = normalize_space(row.get("previousOwner"))
        row["address"] = normalize_space(row.get("address"))

    for idx, deed in enumerate(deeds):
        deed["_idx"] = idx
        deed["lotInt"] = parse_lot(deed.get("lot"))
        deed["deedDateObj"] = parse_us_date(deed.get("normalizedDate"))
        deed["grantor"] = normalize_space(deed.get("grantor"))
        deed["grantee"] = normalize_space(deed.get("grantee"))

    page_maps: dict[str, dict[int, str]] = {}
    for source_key, filename in SOURCE_FILES.items():
        src_path = data_dir / filename
        if src_path.exists():
            page_maps[source_key] = parse_pages(src_path)

    repaired_dates = 0
    for deed in deeds:
        source = deed.get("source")
        page = deed.get("page")
        if source not in page_maps or not isinstance(page, int):
            continue
        pages = page_maps[source]
        chunk = "\n".join(pages.get(p, "") for p in range(page, page + 4))
        primary = "\n".join(pages.get(p, "") for p in range(page, page + 2))
        extracted_dated = extract_dated_line(chunk)
        recording_stamp = extract_recording_stamp(primary)

        current = deed.get("deedDateObj")
        replacement = None
        reason = ""

        if current is None and extracted_dated:
            replacement = extracted_dated
            reason = "Filled missing deed date from deed text."
        elif current and recording_stamp and abs((current - recording_stamp).days) > 365:
            if extracted_dated and abs((extracted_dated - recording_stamp).days) < abs((current - recording_stamp).days):
                replacement = extracted_dated
                reason = "Corrected suspicious deed date using deed text (closer to recording stamp)."
            elif extracted_dated is None:
                replacement = recording_stamp
                reason = "Corrected suspicious deed date using recording stamp fallback."

        if replacement and replacement != current:
            deed["deedDateObj"] = replacement
            deed["normalizedDate"] = date_to_str(replacement)
            deed["isPostAmendment"] = replacement >= AMENDMENT_DATE
            deed["dateRepairNote"] = reason
            repaired_dates += 1

    deed_to_resale = match_by_lot(
        left_items=deeds,
        right_items=resale_entries,
        left_date_key="deedDateObj",
        right_date_key="escrowDateObj",
    )

    for deed in deeds:
        notes: list[str] = []
        resale_match = deed_to_resale[deed["_idx"]]
        matched_resale = None
        if resale_match.matched_index is not None and (resale_match.delta_days or 0) <= 365:
            matched_resale = resale_entries[resale_match.matched_index]
            deed["matchedResale"] = matched_resale
            deed["resaleDeltaDays"] = resale_match.delta_days
        else:
            deed["matchedResale"] = None
            deed["resaleDeltaDays"] = None

        if deed["deedDateObj"] is None and matched_resale and matched_resale.get("escrowDateObj"):
            deed["deedDateObj"] = matched_resale["escrowDateObj"]
            deed["normalizedDate"] = matched_resale["escrowDate"]
            notes.append("Filled missing deed date from resale escrow date.")

        if matched_resale and is_bad_name(deed.get("grantor")) and matched_resale.get("previousOwner"):
            deed["grantor"] = matched_resale["previousOwner"]
            notes.append("Filled grantor from resale previous owner.")
        if matched_resale and is_bad_name(deed.get("grantee")) and matched_resale.get("newOwner"):
            deed["grantee"] = matched_resale["newOwner"]
            notes.append("Filled grantee from resale new owner.")

        deed["fillNotes"] = notes

    for idx, row in enumerate(gl_entries):
        row["_idx"] = idx
        row["glDateObj"] = parse_us_date(row.get("date"))
        row["unitInt"] = infer_gl_unit(row)
        row["description"] = normalize_space(row.get("description"))
        row["isPostAmendment"] = bool(row["glDateObj"] and row["glDateObj"] >= AMENDMENT_DATE)

    all_dated_deeds = [d for d in deeds if d.get("deedDateObj")]
    post_deeds = [d for d in all_dated_deeds if d["deedDateObj"] >= AMENDMENT_DATE]
    post_deeds.sort(key=lambda x: (x["deedDateObj"], x.get("lotInt") or 0, x.get("_idx")))

    post_collections = [
        g
        for g in gl_entries
        if g.get("isPostAmendment") and g.get("type") == "collection" and g.get("glDateObj")
    ]
    post_reversals = [
        g
        for g in gl_entries
        if g.get("isPostAmendment") and g.get("type") == "reversal" and g.get("glDateObj")
    ]

    used_collection_ids: set[int] = set()
    deed_to_collection: dict[int, int] = {}
    # Match GL collections to deeds by lot/date even when the deed predates amendment.
    # This keeps early HOA charges visible as full deed rows instead of generic unmatched GL entries.
    match_candidate_deeds = sorted(
        all_dated_deeds,
        key=lambda x: (
            x["deedDateObj"] < AMENDMENT_DATE,
            x["deedDateObj"],
            x.get("lotInt") or 0,
            x.get("_idx"),
        ),
    )
    for deed in match_candidate_deeds:
        lot = deed.get("lotInt")
        if lot is None:
            continue
        target_date = deed["deedDateObj"]
        matched_resale = deed.get("matchedResale")
        if matched_resale and matched_resale.get("escrowDateObj"):
            target_date = matched_resale["escrowDateObj"]

        candidates = []
        for gl in post_collections:
            if gl["_idx"] in used_collection_ids:
                continue
            if gl.get("unitInt") != lot:
                continue
            delta_days = abs((gl["glDateObj"] - target_date).days)
            score = delta_days
            if gl.get("amount") != 500:
                score += 120
            if "transfer" not in gl["description"].lower() and "capital" not in gl["description"].lower():
                score += 10
            candidates.append((score, delta_days, gl["_idx"]))

        if not candidates:
            continue
        candidates.sort()
        best_score, best_delta, best_idx = candidates[0]
        max_delta = 400 if deed["deedDateObj"] >= AMENDMENT_DATE else 120
        if best_delta <= max_delta and best_score <= 550:
            used_collection_ids.add(best_idx)
            deed_to_collection[deed["_idx"]] = best_idx

    used_reversal_ids: set[int] = set()
    deed_reversal_ids: dict[int, list[int]] = defaultdict(list)
    for deed in match_candidate_deeds:
        primary_idx = deed_to_collection.get(deed["_idx"])
        if primary_idx is None:
            continue
        primary = gl_entries[primary_idx]
        lot = deed.get("lotInt")
        if lot is None:
            continue
        for rev in post_reversals:
            if rev["_idx"] in used_reversal_ids:
                continue
            if rev.get("unitInt") != lot:
                continue
            if abs((rev["glDateObj"] - primary["glDateObj"]).days) <= 90:
                deed_reversal_ids[deed["_idx"]].append(rev["_idx"])
                used_reversal_ids.add(rev["_idx"])

    deed_extra_collection_ids: dict[int, list[int]] = defaultdict(list)
    for deed in match_candidate_deeds:
        if deed["deedDateObj"] < AMENDMENT_DATE:
            continue
        primary_idx = deed_to_collection.get(deed["_idx"])
        if primary_idx is None:
            continue
        lot = deed.get("lotInt")
        if lot is None:
            continue

        lotus_grantor = is_lotus_house(deed.get("grantor"))
        expected = 0 if lotus_grantor else 500
        primary = gl_entries[primary_idx]
        current_net = float(primary.get("amount", 0))
        for rev_idx in deed_reversal_ids.get(deed["_idx"], []):
            current_net += float(gl_entries[rev_idx].get("amount", 0))

        extra_candidates = []
        for gl in post_collections:
            if gl["_idx"] in used_collection_ids:
                continue
            if gl.get("unitInt") != lot:
                continue
            if abs((gl["glDateObj"] - primary["glDateObj"]).days) > 30:
                continue
            extra_candidates.append(gl)
        extra_candidates.sort(
            key=lambda g: (
                abs((g["glDateObj"] - primary["glDateObj"]).days),
                abs(float(g.get("amount", 0))),
                g["_idx"],
            )
        )

        for gl in extra_candidates:
            proposal = current_net + float(gl.get("amount", 0))
            if abs(proposal - expected) < abs(current_net - expected):
                deed_extra_collection_ids[deed["_idx"]].append(gl["_idx"])
                used_collection_ids.add(gl["_idx"])
                current_net = proposal

    analysis_deeds = [
        d
        for d in all_dated_deeds
        if d["deedDateObj"] >= AMENDMENT_DATE
        or d["_idx"] in deed_to_collection
        or d["_idx"] in deed_reversal_ids
    ]
    analysis_deeds.sort(key=lambda x: (x["deedDateObj"], x.get("lotInt") or 0, x.get("_idx")))

    cross_rows: list[dict[str, Any]] = []
    for deed in analysis_deeds:
        lot = deed.get("lotInt")
        grantor = deed.get("grantor") or ""
        grantee = deed.get("grantee") or ""
        lotus_grantor = is_lotus_house(grantor)
        lotus_grantee = is_lotus_house(grantee)
        is_post_amendment = deed["deedDateObj"] >= AMENDMENT_DATE
        due = (not lotus_grantor) if is_post_amendment else None
        expected = (500 if due else 0) if is_post_amendment else None

        primary_gl = gl_entries[deed_to_collection[deed["_idx"]]] if deed["_idx"] in deed_to_collection else None
        primary_amount = float(primary_gl["amount"]) if primary_gl else 0.0
        adjustments = [gl_entries[i] for i in deed_reversal_ids.get(deed["_idx"], [])]
        extra_collections = [gl_entries[i] for i in deed_extra_collection_ids.get(deed["_idx"], [])]
        adjustment_total = sum(float(a["amount"]) for a in adjustments) + sum(
            float(c["amount"]) for c in extra_collections
        )
        actual_net = primary_amount + adjustment_total
        impact = currency(actual_net - expected) if expected is not None else currency(actual_net)

        if not is_post_amendment:
            status = "Pre-Amendment Matched" if (primary_gl or adjustments) else "Pre-Amendment"
        elif actual_net == expected:
            status = "Correct"
        elif actual_net < expected:
            status = "Under-Collected"
        else:
            status = "Over-Collected"

        category = "Conveyance assessment due"
        if not is_post_amendment:
            category = "Pre-amendment transfer with GL activity"
        elif lotus_grantor:
            category = "Exempt transfer (grantor Lotus House)"
        elif lotus_grantee:
            category = "Lotus House as buyer (Claim 1)"

        notes: list[str] = []
        if deed.get("matchedResale"):
            resale = deed["matchedResale"]
            delta = deed.get("resaleDeltaDays")
            notes.append(
                f"Resale escrow {resale['escrowDate']} from {resale['sourceFile']} (delta {delta} days)."
            )
        notes.extend(deed.get("fillNotes", []))
        if deed.get("dateRepairNote"):
            notes.append(deed["dateRepairNote"])
        if primary_gl:
            gl_delta = abs((primary_gl["glDateObj"] - deed["deedDateObj"]).days)
            notes.append(f"GL/deed date delta: {gl_delta} days.")
            if primary_gl.get("amount") != 500:
                notes.append(f"Non-standard collection amount: ${primary_gl['amount']}.")
        if adjustments:
            adj_txt = "; ".join(
                f"{a['date']} ({a['amount']:+.0f})"
                for a in sorted(adjustments, key=lambda x: x["glDateObj"])
            )
            notes.append(f"Included reversal adjustment(s): {adj_txt}.")
        if extra_collections:
            extra_txt = "; ".join(
                f"{c['date']} ({c['amount']:+.0f})"
                for c in sorted(extra_collections, key=lambda x: x["glDateObj"])
            )
            notes.append(f"Included additional same-lot collection(s): {extra_txt}.")
        if due and not primary_gl:
            notes.append("No matching GL collection entry by unit/date.")
        if due is False and primary_gl:
            notes.append("Collection posted on exempt transfer.")
        if not is_post_amendment and primary_gl:
            notes.append("Matched to pre-amendment deed for visibility.")

        cross_rows.append(
            {
                "Phase": deed.get("phase", ""),
                "Lot": lot,
                "Deed Date": deed.get("deedDateObj"),
                "Deed Type": deed.get("deedType", ""),
                "Grantor": grantor,
                "Grantee": grantee,
                "Category": category,
                "Grantor=Lotus House?": "Yes" if lotus_grantor else "No",
                "$500 Due?": ("Yes" if due else "No") if due is not None else "N/A (pre-amendment)",
                "GL Date": primary_gl.get("glDateObj") if primary_gl else None,
                "GL Unit": primary_gl.get("unitInt") if primary_gl else None,
                "GL Description": primary_gl.get("description") if primary_gl else "",
                "Match Status": status,
                "$ Impact": impact,
                "Notes": " ".join(notes),
                "_deedIdx": deed["_idx"],
                "_glIdx": primary_gl.get("_idx") if primary_gl else None,
            }
        )

    matched_gl_ids = set(used_collection_ids) | used_reversal_ids
    unmatched_gl_rows: list[dict[str, Any]] = []
    unmatched_gl_linked_pre = 0
    for gl in sorted(
        [g for g in gl_entries if g.get("isPostAmendment") and g.get("type") in {"collection", "reversal"}],
        key=lambda x: (x["glDateObj"], x["_idx"]),
    ):
        if gl["_idx"] in matched_gl_ids:
            continue
        lot = gl.get("unitInt")
        pre_link_note = ""
        category = "Unmatched GL entry"
        if isinstance(lot, int):
            deed_candidates = [
                d for d in deeds if d.get("lotInt") == lot and isinstance(d.get("deedDateObj"), date)
            ]
            if deed_candidates:
                nearest = min(
                    deed_candidates,
                    key=lambda d: abs((d["deedDateObj"] - gl["glDateObj"]).days),
                )
                delta = abs((nearest["deedDateObj"] - gl["glDateObj"]).days)
                if delta <= 120 and nearest["deedDateObj"] < AMENDMENT_DATE:
                    category = "GL tied to pre-amendment deed"
                    pre_link_note = (
                        f"Nearest deed for lot {lot}: {date_to_str(nearest['deedDateObj'])} "
                        f"(recording #{nearest.get('recordingNumber', '')}), {delta} days from GL."
                    )
                    unmatched_gl_linked_pre += 1
        unmatched_gl_rows.append(
            {
                "Phase": "",
                "Lot": gl.get("unitInt"),
                "Deed Date": None,
                "Deed Type": "",
                "Grantor": "",
                "Grantee": "",
                "Category": category,
                "Grantor=Lotus House?": "",
                "$500 Due?": "",
                "GL Date": gl.get("glDateObj"),
                "GL Unit": gl.get("unitInt"),
                "GL Description": gl.get("description", ""),
                "Match Status": "Unmatched GL",
                "$ Impact": currency(float(gl.get("amount", 0))),
                "Notes": " ".join(
                    part
                    for part in [
                        f"Type={gl.get('type')}; source={gl.get('source')}; original date={gl.get('date')}.",
                        pre_link_note,
                    ]
                    if part
                ),
                "_deedIdx": None,
                "_glIdx": gl.get("_idx"),
            }
        )

    payout_entries = [g for g in gl_entries if g.get("type") in {"payout_to_lotus", "void"}]
    payouts_by_date: dict[date, list[dict[str, Any]]] = defaultdict(list)
    for entry in payout_entries:
        if entry.get("glDateObj"):
            payouts_by_date[entry["glDateObj"]].append(entry)

    payout_descriptions = {
        date(2021, 3, 10): "Transfer Contributions Fees",
        date(2022, 1, 25): "Payment per Agreement (void + re-entry, net $14,500)",
        date(2023, 3, 3): "Payment per agreement",
        date(2024, 3, 8): "Payment per Agreement",
        date(2025, 9, 19): "Payment per agreement",
    }

    lotus_payments: list[dict[str, Any]] = []
    for payout_date, entries in sorted(payouts_by_date.items()):
        net_outflow = -sum(float(e["amount"]) for e in entries)
        if net_outflow <= 0:
            continue
        desc = payout_descriptions.get(
            payout_date,
            " / ".join(
                sorted(
                    {
                        normalize_space(
                            re.sub(r"^Lotus House Development Corp.*?;\s*", "", e["description"])
                        )
                        for e in entries
                    }
                )
            ),
        )
        source_detail = " | ".join(
            f"{e['type']} {e['amount']:+.0f}: {e['description']}" for e in entries
        )
        lotus_payments.append(
            {
                "date": payout_date,
                "amount": currency(net_outflow),
                "description": desc,
                "sourceDetail": source_detail,
            }
        )

    claim1_rows: list[dict[str, Any]] = []
    for row in cross_rows:
        if row["Category"] != "Lotus House as buyer (Claim 1)":
            continue
        claim1_rows.append(
            {
                "date": row["Deed Date"],
                "lot": row["Lot"],
                "grantor": row["Grantor"],
                "grantee": row["Grantee"],
                "amount": 500.0,
                "basis": "Lotus House as grantee on post-amendment transfer.",
                "notes": row["Notes"],
            }
        )
    claim1_rows.sort(key=lambda x: (x["date"], x["lot"]))
    claim1_total = currency(sum(x["amount"] for x in claim1_rows))
    claim2_total = currency(sum(x["amount"] for x in lotus_payments))
    grand_total = currency(claim1_total + claim2_total)

    lot_reference_rows: list[dict[str, Any]] = []
    lots: set[int] = set()
    lots.update(d["lotInt"] for d in deeds if isinstance(d.get("lotInt"), int) and d["lotInt"] > 0)
    lots.update(r["lotInt"] for r in resale_entries if isinstance(r.get("lotInt"), int) and r["lotInt"] > 0)
    lots.update(g["unitInt"] for g in gl_entries if isinstance(g.get("unitInt"), int) and g["unitInt"] > 0)

    deeds_by_lot: dict[int, list[dict[str, Any]]] = defaultdict(list)
    resale_by_lot: dict[int, list[dict[str, Any]]] = defaultdict(list)
    gl_by_lot: dict[int, list[dict[str, Any]]] = defaultdict(list)

    for d in deeds:
        lot = d.get("lotInt")
        if isinstance(lot, int) and lot > 0:
            deeds_by_lot[lot].append(d)
    for r in resale_entries:
        lot = r.get("lotInt")
        if isinstance(lot, int) and lot > 0:
            resale_by_lot[lot].append(r)
    for g in gl_entries:
        lot = g.get("unitInt")
        if isinstance(lot, int) and lot > 0:
            gl_by_lot[lot].append(g)

    for lot in sorted(lots):
        lot_deeds = deeds_by_lot.get(lot, [])
        lot_resale = sorted(resale_by_lot.get(lot, []), key=lambda x: x.get("escrowDateObj") or date.min)
        lot_gl = [g for g in gl_by_lot.get(lot, []) if g.get("isPostAmendment") and g.get("type") in {"collection", "reversal"}]
        phases = sorted({normalize_space(d.get("phase")) for d in lot_deeds if d.get("phase")})
        latest_resale = lot_resale[-1] if lot_resale else None
        latest_address = ""
        if latest_resale and latest_resale.get("address"):
            latest_address = latest_resale["address"]
        elif lot_resale:
            latest_address = next((x["address"] for x in reversed(lot_resale) if x.get("address")), "")

        sources = []
        if lot_deeds:
            sources.append("deeds")
        if lot_resale:
            sources.append("resale")
        if lot_gl:
            sources.append("gl")

        lot_reference_rows.append(
            {
                "Lot": lot,
                "Phase(s)": ", ".join(phases),
                "Address": latest_address,
                "Latest Owner (Resale)": latest_resale.get("newOwner", "") if latest_resale else "",
                "Previous Owner (Resale)": latest_resale.get("previousOwner", "") if latest_resale else "",
                "Account #": latest_resale.get("accountNumber", "") if latest_resale else "",
                "Latest Escrow Date": latest_resale.get("escrowDateObj") if latest_resale else None,
                "Latest Process Date": latest_resale.get("processDateObj") if latest_resale else None,
                "Deed Count": len(lot_deeds),
                "Post-Amendment Deed Count": sum(
                    1
                    for d in lot_deeds
                    if d.get("deedDateObj") and d["deedDateObj"] >= AMENDMENT_DATE
                ),
                "GL Entry Count (Post-Amendment)": len(lot_gl),
                "GL Net Amount (Post-Amendment)": currency(
                    sum(float(g.get("amount", 0)) for g in lot_gl)
                ),
                "Sources": ", ".join(sources),
            }
        )

    resale_missing_deed: list[dict[str, Any]] = []
    deed_by_lot: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for d in deeds:
        lot = d.get("lotInt")
        if isinstance(lot, int) and lot > 0 and d.get("deedDateObj"):
            deed_by_lot[lot].append(d)

    for lot, lot_deeds in deed_by_lot.items():
        lot_deeds.sort(key=lambda x: x["deedDateObj"])

    for resale in resale_entries:
        lot = resale.get("lotInt")
        escrow = resale.get("escrowDateObj")
        if not isinstance(lot, int) or lot <= 0 or not isinstance(escrow, date):
            continue
        candidates = deed_by_lot.get(lot, [])
        if not candidates:
            resale_missing_deed.append(
                {
                    "lot": lot,
                    "escrowDate": escrow,
                    "newOwner": resale.get("newOwner", ""),
                    "address": resale.get("address", ""),
                    "reason": "No deed found for lot in deed dataset.",
                }
            )
            continue
        nearest = min(candidates, key=lambda d: abs((d["deedDateObj"] - escrow).days))
        nearest_delta = abs((nearest["deedDateObj"] - escrow).days)
        if nearest_delta > 180:
            resale_missing_deed.append(
                {
                    "lot": lot,
                    "escrowDate": escrow,
                    "newOwner": resale.get("newOwner", ""),
                    "address": resale.get("address", ""),
                    "reason": f"Nearest deed date {nearest['normalizedDate']} is {nearest_delta} days away.",
                }
            )

    workbook_path = project_root / "conveyance_assessment_analysis.xlsx"
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Deed-GL Cross Reference"
    headers1 = [
        "Phase",
        "Lot",
        "Deed Date",
        "Deed Type",
        "Grantor",
        "Grantee",
        "Category",
        "Grantor=Lotus House?",
        "$500 Due?",
        "GL Date",
        "GL Unit",
        "GL Description",
        "Match Status",
        "$ Impact",
        "Notes",
    ]
    ws1.append(headers1)

    ordered_cross = sorted(cross_rows, key=lambda r: (r["Deed Date"] or date.min, r["Lot"] or 0))
    ordered_cross.extend(unmatched_gl_rows)

    for row in ordered_cross:
        ws1.append([row[h] for h in headers1])

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    orange_fill = PatternFill("solid", fgColor="FCE4D6")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    blue_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_idx in range(2, ws1.max_row + 1):
        status = ws1.cell(row=row_idx, column=13).value
        fill = None
        if status == "Correct":
            fill = green_fill
        elif status == "Under-Collected":
            fill = red_fill
        elif status == "Over-Collected":
            fill = orange_fill
        elif status == "Pre-Amendment Matched":
            fill = blue_fill
        elif status == "Unmatched GL":
            fill = yellow_fill
        if fill:
            for col_idx in range(1, ws1.max_column + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = fill

        ws1.cell(row=row_idx, column=3).number_format = "mm/dd/yyyy"
        ws1.cell(row=row_idx, column=10).number_format = "mm/dd/yyyy"
        ws1.cell(row=row_idx, column=14).number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'
        ws1.cell(row=row_idx, column=15).alignment = Alignment(wrap_text=True, vertical="top")

    ws1.freeze_panes = "A2"
    auto_fit_columns(ws1)

    ws2 = wb.create_sheet("Lotus House Payments")
    headers2 = ["Date", "Amount", "Description", "Source GL Entries"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for item in lotus_payments:
        ws2.append([item["date"], item["amount"], item["description"], item["sourceDetail"]])
    ws2.append(["TOTAL", currency(sum(x["amount"] for x in lotus_payments)), "", ""])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    ws2.cell(row=ws2.max_row, column=2).font = Font(bold=True)
    for row_idx in range(2, ws2.max_row + 1):
        ws2.cell(row=row_idx, column=1).number_format = "mm/dd/yyyy"
        ws2.cell(row=row_idx, column=2).number_format = '"$"#,##0.00'
        ws2.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    auto_fit_columns(ws2)

    ws3 = wb.create_sheet("Lotus House Claim Detail")
    headers3 = ["Claim", "Date", "Lot", "Grantor", "Grantee", "Amount", "Basis", "Notes"]
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for item in claim1_rows:
        ws3.append(
            [
                "Claim 1",
                item["date"],
                item["lot"],
                item["grantor"],
                item["grantee"],
                item["amount"],
                item["basis"],
                item["notes"],
            ]
        )
    for payment in lotus_payments:
        ws3.append(
            [
                "Claim 2",
                payment["date"],
                "",
                "",
                "Lotus House Development Corp",
                payment["amount"],
                "Improper transfer from HOA fund to Lotus House.",
                payment["description"],
            ]
        )
    ws3.append(["", "", "", "", "", "", "", ""])
    ws3.append(["Claim 1 Subtotal", "", "", "", "", claim1_total, "", ""])
    ws3.append(["Claim 2 Subtotal", "", "", "", "", claim2_total, "", ""])
    ws3.append(["Grand Total", "", "", "", "", grand_total, "", ""])
    for row_idx in range(2, ws3.max_row + 1):
        ws3.cell(row=row_idx, column=2).number_format = "mm/dd/yyyy"
        ws3.cell(row=row_idx, column=6).number_format = '"$"#,##0.00'
        ws3.cell(row=row_idx, column=8).alignment = Alignment(wrap_text=True, vertical="top")
    for row_idx in range(ws3.max_row - 2, ws3.max_row + 1):
        ws3.cell(row=row_idx, column=1).font = Font(bold=True)
        ws3.cell(row=row_idx, column=6).font = Font(bold=True)
    auto_fit_columns(ws3)

    ws4 = wb.create_sheet("Lot Reference")
    headers4 = [
        "Lot",
        "Phase(s)",
        "Address",
        "Latest Owner (Resale)",
        "Previous Owner (Resale)",
        "Account #",
        "Latest Escrow Date",
        "Latest Process Date",
        "Deed Count",
        "Post-Amendment Deed Count",
        "GL Entry Count (Post-Amendment)",
        "GL Net Amount (Post-Amendment)",
        "Sources",
    ]
    ws4.append(headers4)
    for cell in ws4[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for item in lot_reference_rows:
        ws4.append([item[h] for h in headers4])
    for row_idx in range(2, ws4.max_row + 1):
        ws4.cell(row=row_idx, column=7).number_format = "mm/dd/yyyy"
        ws4.cell(row=row_idx, column=8).number_format = "mm/dd/yyyy"
        ws4.cell(row=row_idx, column=12).number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'
    ws4.freeze_panes = "A2"
    auto_fit_columns(ws4)

    wb.save(workbook_path)

    post_scope_rows = [r for r in cross_rows if isinstance(r["Deed Date"], date) and r["Deed Date"] >= AMENDMENT_DATE]
    due_rows = [r for r in post_scope_rows if r["$500 Due?"] == "Yes"]
    exempt_rows = [r for r in post_scope_rows if r["$500 Due?"] == "No"]
    pre_rows = [r for r in cross_rows if r["$500 Due?"] == "N/A (pre-amendment)"]
    claim1_count = len(claim1_rows)
    status_counts: dict[str, int] = defaultdict(int)
    for row in cross_rows:
        status_counts[row["Match Status"]] += 1

    spot_candidates = [
        r for r in cross_rows if r["Match Status"] == "Correct" and r["GL Date"] and r["Deed Date"]
    ]
    spot_candidates.sort(key=lambda x: abs((x["GL Date"] - x["Deed Date"]).days))
    spot_checks = spot_candidates[:10]

    report_path = working_dir / "verification_report.md"
    lines = []
    lines.append("# Verification Report")
    lines.append("")
    lines.append("## Data Validation")
    lines.append(f"- Deeds loaded: {len(deeds)}")
    lines.append(f"- Deed dates auto-corrected from source pages: {repaired_dates}")
    lines.append(f"- Post-amendment deeds: {len(post_deeds)}")
    lines.append(f"- Deeds missing normalized date after recovery: {sum(1 for d in deeds if not d.get('deedDateObj'))}")
    lines.append(f"- Resale rows parsed: {len(resale_entries)}")
    lines.append(
        f"- Resale rows with unparseable escrow date: {sum(1 for r in resale_entries if not r.get('escrowDateObj'))}"
    )
    lines.append(f"- GL rows loaded: {len(gl_entries)}")
    lines.append(
        f"- Post-amendment GL collection/reversal rows: {len([g for g in gl_entries if g.get('isPostAmendment') and g.get('type') in {'collection', 'reversal'}])}"
    )
    lines.append("")
    lines.append("## Match Summary")
    lines.append(f"- Post-amendment rows in scope: {len(post_scope_rows)}")
    lines.append(f"- Due transfers (`$500 Due? = Yes`): {len(due_rows)}")
    lines.append(f"- Exempt transfers (`grantor = Lotus House`): {len(exempt_rows)}")
    lines.append(f"- Pre-amendment matched rows included for context: {len(pre_rows)}")
    lines.append(f"- Correct rows: {status_counts.get('Correct', 0)}")
    lines.append(f"- Under-collected rows: {status_counts.get('Under-Collected', 0)}")
    lines.append(f"- Over-collected rows: {status_counts.get('Over-Collected', 0)}")
    lines.append(f"- Pre-amendment matched status rows: {status_counts.get('Pre-Amendment Matched', 0)}")
    lines.append(f"- Unmatched GL rows: {len(unmatched_gl_rows)}")
    lines.append(f"- Unmatched GL rows tied to pre-amendment deeds: {unmatched_gl_linked_pre}")
    lines.append("")
    lines.append("## Spot-Check Matches")
    if not spot_checks:
        lines.append("- No matched rows available for spot-check.")
    else:
        for row in spot_checks:
            lines.append(
                f"- Lot {row['Lot']}: deed {row['Deed Date'].strftime('%m/%d/%Y')} vs GL {row['GL Date'].strftime('%m/%d/%Y')} ({abs((row['GL Date'] - row['Deed Date']).days)} days), status={row['Match Status']}."
            )
    lines.append("")
    lines.append("## Resale Entries Missing Deed Coverage")
    lines.append(f"- Count: {len(resale_missing_deed)}")
    for item in resale_missing_deed[:20]:
        lines.append(
            f"- Lot {item['lot']} escrow {item['escrowDate'].strftime('%m/%d/%Y')}: {item['newOwner']} ({item['reason']})"
        )
    if len(resale_missing_deed) > 20:
        lines.append(f"- ... plus {len(resale_missing_deed) - 20} additional rows.")
    lines.append("")
    lines.append("## Claim Totals")
    lines.append(f"- Claim 1 count: {claim1_count} transfers")
    lines.append(f"- Claim 1 subtotal: ${claim1_total:,.2f}")
    lines.append(f"- Claim 2 subtotal: ${claim2_total:,.2f}")
    lines.append(f"- Grand total: ${grand_total:,.2f}")

    report_path.write_text("\n".join(lines), encoding="utf-8")

    summary_path = working_dir / "analysis_summary.json"
    summary = {
        "workbook": str(workbook_path),
        "postAmendmentDeeds": len(post_deeds),
        "crossRows": len(cross_rows),
        "statusCounts": dict(status_counts),
        "unmatchedGlRows": len(unmatched_gl_rows),
        "resaleRows": len(resale_entries),
        "resaleMissingDeedRows": len(resale_missing_deed),
        "repairedDeedDates": repaired_dates,
        "claim1Count": claim1_count,
        "claim1Subtotal": claim1_total,
        "claim2Subtotal": claim2_total,
        "grandTotal": grand_total,
        "unmatchedGlLinkedPreAmendment": unmatched_gl_linked_pre,
    }
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    resale_out = working_dir / "resale_parsed.json"
    resale_serializable = []
    for r in resale_entries:
        copy = dict(r)
        copy.pop("_idx", None)
        copy["escrowDateObj"] = copy["escrowDateObj"].strftime("%m/%d/%Y") if copy.get("escrowDateObj") else ""
        copy["processDateObj"] = copy["processDateObj"].strftime("%m/%d/%Y") if copy.get("processDateObj") else ""
        resale_serializable.append(copy)
    resale_out.write_text(json.dumps(resale_serializable, indent=2), encoding="utf-8")

    print(f"Wrote workbook: {workbook_path}")
    print(f"Wrote verification report: {report_path}")
    print(f"Wrote summary: {summary_path}")


if __name__ == "__main__":
    main()
